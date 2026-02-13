#!/usr/bin/env python3
"""
Script to read CZ_MasterSheet.xlsx and analyze its structure
"""

try:
    import pandas as pd
    import openpyxl
    
    # Try to read the Excel file
    try:
        xl_file = pd.ExcelFile('CZ_MasterSheet.xlsx')
        print('Available sheets:', xl_file.sheet_names)
        print()
        
        # Read each sheet to see its structure
        for sheet_name in xl_file.sheet_names:
            print(f'Sheet: {sheet_name}')
            df = pd.read_excel('CZ_MasterSheet.xlsx', sheet_name=sheet_name)
            print(f'Columns: {list(df.columns)}')
            print(f'Rows: {len(df)}')
            print('Sample data (first 3 rows):')
            if not df.empty:
                print(df.head(3).to_string(max_cols=10))
            print('=' * 80)
            print()
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        
except ImportError as e:
    print(f"Required modules not available: {e}")
    print("Please install: pip install pandas openpyxl")
    
    # Fallback: try with openpyxl only
    try:
        import openpyxl
        wb = openpyxl.load_workbook('CZ_MasterSheet.xlsx', read_only=True)
        print("Available sheets (openpyxl):", wb.sheetnames)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Sheet: {sheet_name}")
            print(f"Dimensions: {ws.max_row} rows, {ws.max_column} columns")
            
            # Get headers (first row)
            if ws.max_row > 0:
                headers = []
                for col in range(1, min(ws.max_column + 1, 21)):  # Limit to first 20 columns
                    cell_value = ws.cell(row=1, column=col).value
                    if cell_value:
                        headers.append(str(cell_value))
                    else:
                        headers.append(f"Col{col}")
                print(f"Headers: {headers}")
            print("=" * 80)
            print()
    except ImportError:
        print("Neither pandas nor openpyxl available. Cannot read Excel file.")
    except Exception as e:
        print(f"Error reading with openpyxl: {e}")